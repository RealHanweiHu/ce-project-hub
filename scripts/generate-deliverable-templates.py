#!/usr/bin/env python3
"""交付物 Excel 模板生成器。

从 shared/sop-templates.ts + shared/task-deliverables.ts 导出的交付物词表
（scripts 同级运行 `npx tsx` 导出，或直接读缓存 tsv），按 14 个模板族为
每个交付物名称生成一份定制封面的 .xlsx 模板，输出到
docs/templates/deliverables/<F##-族名>/<交付物名>.xlsx，并生成总索引。

依赖：python3 + openpyxl（pip install openpyxl）。
用法：python3 scripts/generate-deliverable-templates.py [deliverables.tsv]
tsv 格式：首行 TOTAL 行忽略；其后每行 `名称\tcat:P#,cat:P#`。
"""
import os
import re
import sys
import subprocess
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_ROOT = os.path.join(ROOT, "docs", "templates", "deliverables")

# ── 设计基调（与交付物模板库 Artifact 一致：制图蓝 + 冷灰） ──────────────
FONT_NAME = "Microsoft YaHei"
C_ACCENT = "2C5FA8"      # 制图蓝
C_ACCENT_SOFT = "E8EEF7"
C_LINE = "B8C2CC"
C_MUTED = "74808C"
C_REQ = "B45309"         # 必填橙

F_TITLE = Font(name=FONT_NAME, size=18, bold=True, color="1B232E")
F_H = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
F_SEC = Font(name=FONT_NAME, size=11, bold=True, color="1B232E")
F_BODY = Font(name=FONT_NAME, size=10, color="1B232E")
F_HINT = Font(name=FONT_NAME, size=9, italic=True, color=C_MUTED)
F_TAG = Font(name=FONT_NAME, size=9, bold=True, color=C_ACCENT)
F_REQ = Font(name=FONT_NAME, size=10, bold=True, color=C_REQ)

FILL_H = PatternFill("solid", start_color=C_ACCENT)
FILL_SEC = PatternFill("solid", start_color=C_ACCENT_SOFT)
THIN = Side(style="thin", color=C_LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

FAMILIES = {
    "F01": "评审记录表",
    "F02": "测试报告",
    "F03": "问题与跟踪清单",
    "F04": "影响分析与评估",
    "F05": "BOM与物料",
    "F06": "计划与排程",
    "F07": "文件与清单",
    "F08": "试产与良率报告",
    "F09": "SOP·WI与检验标准",
    "F10": "认证证据归档",
    "F11": "复用确认单",
    "F12": "客户签核单",
    "F13": "需求与文档",
    "F14": "样机与Build记录",
}

# 明确指派（优先于规则）
OVERRIDES = {
    "商业可行性分析": "F13",
    "专利检索与规避分析": "F13",
    "核心技术 POC 报告": "F13",
    "电芯规格书与既有认证资料": "F10",
    "文件与认证资料发布记录": "F10",
    "模块复用策略与影响分析": "F04",
    "安全FMEA与危害分析": "F04",
    "DFMEA/风险清单": "F04",
    "PFMEA/CTQ控制计划": "F04",
    "报价": "F13",
    "样品与报价": "F13",
    "RACI 责任矩阵": "F13",
    "供应商评估表": "F07",
    "量产测试设备清单": "F07",
    "资源与供应商确认清单": "F07",
    "EOL 100%测试项目清单": "F07",
    "关键安全/性能检验标准": "F09",
    "限度样本与检验标准": "F09",
    "限度样本": "F14",
    "治具与测试程序": "F14",
    "测试程序与治具": "F14",
    "治具/EOL测试程序验收": "F02",
    "EOL 100%测试能力验收记录": "F02",
    "版本切换与库存处理方案": "F06",
    "库存处理方案": "F06",
    "版本切换计划": "F06",
    "产线切换计划": "F06",
    "关键路径标注": "F06",
    "Kickoff 会议纪要": "F01",
    "IDR Kickoff 记录": "F01",
    "CCB 变更决策记录": "F01",
    "ECN 正式发布": "F10",
    "ECN/ECR 记录": "F03",
    "ECN/ECR记录": "F03",
    "工程变更 ECN 记录": "F03",
    "ECR变更申请书": "F13",
    "变更需求单 ECR": "F13",
    "用户访谈纪要": "F13",
    "用户痛点 Top10": "F13",
    "用户旅程地图": "F13",
    "竞品对比矩阵": "F13",
    "销量预测模型": "F13",
    "目标成本与定价分析": "F13",
    "现有设计基线包": "F07",
    "翻新需求与边界定义": "F13",
    "IDR翻新 brief": "F13",
    "迭代需求书": "F13",
    "立项申请书": "F13",
    "市场调研报告": "F13",
    "产品概念书": "F13",
    "客户特殊要求清单 CSR": "F13",
    "售后数据分析": "F13",
    "售后问题报表": "F03",
    "售后问题跟踪": "F03",
    "售后影响跟踪报表": "F03",
    "市场与售后影响跟踪报表": "F03",
    "良率周报": "F08",
    "良率监控报表": "F08",
    "量产工艺评估报告": "F04",
    "模具可行性评估": "F04",
    "可行性与初步 DFM 报告": "F04",
    "DFM/可制造性反馈报告": "F04",
    "DFM 变更评审报告": "F01",
    "DFM/DFA 评审报告": "F01",
    "DFM/DFT/DFMEA评审记录": "F01",
    "保护参数与热路径校核": "F02",
    "新旧版本对比报告": "F02",
    "openBOM 核对清单": "F05",
    "图纸/规格完整性确认": "F07",
    "保护电路设计评审或复用确认": "F11",
    "电芯厂质量审核或复用资质确认": "F11",
    "电芯复用/定点与二供策略": "F11",
    "认证补测/复用确认": "F11",
    "UN38.3运输测试报告或复用确认": "F10",
    "电芯/电池包安全认证报告或复用确认": "F10",
    "量产产品": "F14",
    "试产/首批量产产品": "F14",
    "试产50-300台": "F14",
    "安全验证项目与判定标准": "F09",
    "外观检验标准": "F09",
    "认证路线图": "F06",
    "认证路线图初判": "F06",
    "认证路径预判": "F06",
    "认证前置依赖与送样计划": "F06",
    "版本/接口/OTA回归范围": "F07",
    "性能/寿命设计边界": "F07",
    "装配/公差/材料方案": "F07",
    "保护电路设计输入": "F07",
    "初步 BOM 与 NRE/模具方案": "F05",
    "模具/治具归属与 NRE 确认": "F13",
    "模具费用与周期确认": "F13",
    "安全件供应商资质清单": "F07",
    "电芯供应风险清单": "F03",
    "长交期/单一来源料件清单": "F05",
    "关键技术挑战清单": "F03",
    "关键料件规格确认": "F05",
    "安全件批准版本记录": "F05",
    "电芯供应商整改闭环记录": "F03",
    "PVT Readiness清单": "F07",
    "DVT输入冻结记录": "F01",
    "EVT可选项裁剪记录": "F07",
    "CIP 持续改善报告": "F13",
    "产能爬坡计划": "F06",
    "市场上市计划": "F06",
    "更新后的SOP/WI": "F09",
    "CMF 方案": "F07",
    "整机功能与兼容回归报告": "F02",
    "包装与物流验证": "F02",
    "PCB v2": "F07",
    "PCB v2 改板记录": "F07",
    "T0/T1试模准备记录": "F14",
    "变更设计评审报告": "F01",
    "PVT或首批质量报告": "F08",
    "FAI 首件检验报告": "F02",
    "投模评审/开模批准记录": "F01",
    "版本文件发布记录": "F10",
    "首批量产稳定性报告": "F08",
    "市场与渠道反馈清单": "F03",
    "版本资料移交清单": "F07",
    "IDR项目关闭报告": "F01",
    "产品规格基线确认记录": "F01",
    "六模块执行基线": "F07",
    "风险声明与评估结论": "F04",
    "首批量产与爬坡记录": "F08",
    "稳定期周报与QA结论": "F08",
    "投模评审与开模批准记录": "F01",
    "整机功能与兼容性回归报告": "F02",
    "外观标准与限度样本": "F09",
    "配件确认记录": "F07",
    "试产良率与问题关闭报告": "F08",
    "认证与运输证据覆盖复核记录": "F10",
}

# 规则（按顺序匹配）
PATTERNS = [
    (r"客户|签样|放行记录|规格确认书（客户", "F12"),
    (r"设计输入冻结确认", "F12"),
    (r"评审记录|评审纪要|冻结记录|Kickoff|决策记录|会议纪要", "F01"),
    (r"复用确认|复用资质|二供策略", "F11"),
    (r"问题清单|问题关闭|问题预警|问题报表|问题跟踪", "F03"),
    (r"影响分析|影响评估|影响确认|影响跟踪", "F04"),
    (r"FMEA|危害分析|风险清单", "F04"),
    (r"UN38\.3|MSDS|认证报告|安规|认证更新|重认证|认证资料", "F10"),
    (r"BOM|料件|物料|齐套", "F05"),
    (r"试产|良率|FPY|首批量产报告|小批", "F08"),
    (r"SOP|WI|检验标准|作业指导", "F09"),
    (r"样机|样品|样件|Build Record|首件", "F14"),
    (r"测试报告|试模报告|验证报告|回归测试|联调|测试方案|验收记录|测试能力", "F02"),
    (r"计划|时程|甘特|WBS|里程碑|路线图|策略|方案|切换|库存", "F06"),
    (r"设计包|图纸|外观图|结构图|原理图|Layout|架构|设计稿|设计输入|3D|CMF|设计边界", "F07"),
    (r"清单|矩阵|索引", "F07"),
    (r"报表|报告|记录|数据", "F13"),
]


def classify(name: str) -> str:
    if name in OVERRIDES:
        return OVERRIDES[name]
    for pat, fid in PATTERNS:
        if re.search(pat, name):
            return fid
    return "F13"


def sanitize(name: str) -> str:
    out = name
    for a, b in [("/", "·"), ("\\", "·"), (":", "："), ("*", "×"), ("?", "？"),
                 ('"', "'"), ("<", "("), (">", ")"), ("|", "·")]:
        out = out.replace(a, b)
    return out.strip()


# ── 通用构件 ────────────────────────────────────────────────────────────

def _ws_defaults(ws, widths):
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _cell(ws, ref, value, font=F_BODY, fill=None, align=LEFT, border=True):
    c = ws[ref]
    c.value = value
    c.font = font
    if fill:
        c.fill = fill
    c.alignment = align
    if border:
        c.border = BORDER
    return c


def _merge(ws, ref):
    ws.merge_cells(ref)
    return ws[ref.split(":")[0]]


def _table_header(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_H
        c.fill = FILL_H
        c.alignment = CENTER
        c.border = BORDER


def _blank_rows(ws, start, count, ncols, height=20):
    for r in range(start, start + count):
        ws.row_dimensions[r].height = height
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = F_BODY
            cell.alignment = WRAP


def _dv(ws, options, ref_range):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ref_range)


def _section(ws, row, ncols, text, hint=None):
    _merge(ws, f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SEC
    c.fill = FILL_SEC
    c.alignment = LEFT
    for i in range(1, ncols + 1):
        ws.cell(row=row, column=i).border = BORDER
        ws.cell(row=row, column=i).fill = FILL_SEC
    if hint:
        r = row + 1
        _merge(ws, f"A{r}:{get_column_letter(ncols)}{r}")
        h = ws.cell(row=r, column=1, value=hint)
        h.font = F_HINT
        h.alignment = WRAP
        return r + 1
    return row + 1


def _signoff(ws, row, ncols, roles=("编制", "审核", "批准")):
    row = _section(ws, row, ncols, "签核 Signoff")
    per = max(2, ncols // len(roles))
    col = 1
    for role in roles:
        end = min(col + per - 1, ncols)
        _merge(ws, f"{get_column_letter(col)}{row}:{get_column_letter(end)}{row}")
        c = ws.cell(row=row, column=col, value=f"{role}：")
        c.font = F_BODY
        c.alignment = LEFT
        for i in range(col, end + 1):
            ws.cell(row=row, column=i).border = BORDER
        _merge(ws, f"{get_column_letter(col)}{row + 1}:{get_column_letter(end)}{row + 1}")
        d = ws.cell(row=row + 1, column=col, value="日期：")
        d.font = F_HINT
        for i in range(col, end + 1):
            ws.cell(row=row + 1, column=i).border = BORDER
        col = end + 1
        if col > ncols:
            break
    ws.row_dimensions[row].height = 28
    ws.row_dimensions[row + 1].height = 18
    return row + 2


def _print_setup(ws, last_row, ncols):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:{get_column_letter(ncols)}{last_row}"


# ── v2：内容预置（按交付物名称匹配，打开即用）＋ 结构简化（少封面/少 sheet/给示例行） ──

F_EX = Font(name=FONT_NAME, size=9.5, italic=True, color="9AA6B2")

# 测试项预置（气泵/锂电工厂场景），按名称关键词匹配，命中即为每项建一个 sheet
TEST_ITEM_SETS = [
    (r"可靠性|耐久", ["跌落测试", "高低温存储", "湿热测试", "振动测试", "连续工作温升", "按键开关寿命", "整机老化"]),
    (r"保护功能|保护参数|热路径", ["过充保护", "过放保护", "过流保护", "过温保护", "短路保护"]),
    (r"功能测试|功能与性能", ["按键与模式切换", "充气与工作功能", "压力设定与自动停机", "灯光与显示", "充电功能", "低电量保护提示"]),
    (r"性能测试", ["最大压力", "充气速度与流量", "噪音", "续航", "充电时长", "工作温升"]),
    (r"软件|固件", ["功能回归", "边界与异常处理", "断电复位恢复", "OTA与烧录", "错误码核对", "生产测试接口"]),
    (r"试模|修模", ["外观面检查", "关键尺寸测量", "缩水与变形", "试装配合", "材料与颜色确认"]),
    (r"包装|物流", ["含包装跌落", "振动运输", "堆码测试", "标签与说明书核对"]),
    (r"EOL|治具|测试程序", ["测试项覆盖核对", "GRR 重复性", "连续运行稳定性", "不良品拦截验证"]),
    (r"联调", ["开机自检", "传感器标定", "保护触发联动", "显示与通信"]),
    (r"关键模块", ["升级模块专项验证", "新旧版本对比", "接口与整机适配"]),
    (r"回归|对比", ["核心功能回归", "新旧版本性能对比", "附件接头兼容", "异常场景"]),
    (r"FAI|首件", ["外观检查", "尺寸测量", "装配与功能", "标识核对"]),
]
DEFAULT_TEST_ITEMS = ["主要功能验证", "性能抽测", "异常场景"]


def test_items_for(name):
    for pat, items in TEST_ITEM_SETS:
        if re.search(pat, name):
            return items
    return DEFAULT_TEST_ITEMS


# 文档章节骨架（F13），命中关键词即预置章节
DOC_SECTIONS = [
    (r"PRD|需求文档", [("背景与目标", "为什么做、达到什么目标"), ("目标用户与场景", "谁在什么场景用"), ("功能需求", "逐条列功能点与优先级"), ("非功能需求", "性能/安全/合规"), ("验收标准", "每条需求怎么算通过"), ("本期不做", "明确边界防蔓延")]),
    (r"PSD|规格书", [("产品概述", "定位/型号/配置"), ("硬件规格", "泵体/电机/电池/主控"), ("性能指标", "压力/流量/噪音/续航，量化"), ("安全边界", "锂电/受压腔体/温升的保护与验收指标"), ("接口与配件", "充电口/嘴阀/附件"), ("环境与合规", "工作温度/目标市场认证")]),
    (r"概念书", [("一句话定义", "电梯陈述"), ("核心卖点", "3 个 USP"), ("目标用户与场景", ""), ("竞品小结", "对比矩阵结论"), ("商业测算摘要", "目标成本/售价/毛利")]),
    (r"迭代需求书", [("升级目标与代际定义", "1代→2代改什么、为什么"), ("一代 TOP 问题复盘", "售后/RMA 数据，必改项"), ("核心指标对比", "一代 vs 二代目标值"), ("目标市场与 SKU", ""), ("上市窗口与约束", "")]),
    (r"立项申请", [("项目概述", ""), ("市场与商业价值", ""), ("技术与资源评估", ""), ("风险与应对", ""), ("申请决议", "预算/人力/时间窗口")]),
    (r"市场调研", [("市场规模与趋势", ""), ("竞品分析", "TOP 竞品参数/价格/卖点"), ("用户反馈汇总", "客户/渠道输入"), ("机会与建议", "")]),
    (r"商业可行性|定价", [("销量预测", "3 年模型摘要"), ("成本与定价", "目标 BOM/售价"), ("毛利与回收期", ""), ("敏感性与风险", "")]),
    (r"brief|边界定义|需求单|申请书|ECR", [("背景与动因", ""), ("范围与边界", "改什么、不改什么"), ("目标与验收", ""), ("约束与风险", "")]),
    (r"POC|专利|检索", [("对象与方法", ""), ("结果", ""), ("结论与建议", "可行/规避方案")]),
    (r"关闭报告|总结|复盘|改善报告", [("目标达成情况", "对立项目标逐条对账"), ("关键数据", "良率/成本/周期"), ("经验教训", ""), ("遗留与移交", "")]),
    (r"NRE|归属|报价|费用", [("范围与明细", "逐项列金额"), ("归属与条款", "谁出资/所有权"), ("确认结论", "")]),
    (r"RACI", [("角色分工表", "任务 × 角色，标 R/A/C/I"), ("关键接口人", "")]),
]
GENERIC_DOC_SECTIONS = [("背景", ""), ("内容", ""), ("结论", ""), ("附件清单", "编号+存放位置")]


def doc_sections_for(name):
    for pat, secs in DOC_SECTIONS:
        if re.search(pat, name):
            return secs
    return GENERIC_DOC_SECTIONS


# 清单条目预置（F07 设计包/资料包类）
LIST_ITEM_SETS = [
    (r"结构设计包|结构 3D|MD", ["3D 总装模型", "2D 关键零件图", "公差分析", "材料与表面处理清单", "装配指引"]),
    (r"PCBA|原理图|PCB|电子", ["原理图", "PCB Layout", "关键器件清单", "测试点清单", "DFM 检查记录"]),
    (r"软件|架构|SW|OTA", ["架构/流程图", "版本与接口清单", "回归范围说明", "烧录与产测接口"]),
    (r"电池|电源", ["电芯/电池包规格书", "保护板设计资料", "充电策略说明", "热路径与温升校核"]),
    (r"机芯|泵体", ["机芯图纸/规格", "性能曲线", "寿命与噪音数据", "装配接口说明"]),
    (r"ID|CMF|外观", ["ID 效果图", "CMF 色板与工艺", "丝印/标识稿", "关键外观面定义"]),
    (r"基线包", ["BOM", "图纸包", "软件版本", "测试报告", "认证资料", "售后/RMA 数据"]),
    (r"包装|标签", ["包装结构图", "彩盒/说明书稿", "标签与铭牌稿", "运输标识"]),
    (r"Readiness|就绪", ["物料齐套", "治具/测试程序", "SOP/WI", "人员培训", "关键风险关闭"]),
    (r"设计输入", ["客户输入文件", "规格要求", "接口定义", "验收口径"]),
]


def list_items_for(name):
    for pat, items in LIST_ITEM_SETS:
        if re.search(pat, name):
            return items
    return []


# 计划行预置（F06）
PLAN_ROW_SETS = [
    (r"迭代项目计划", [("P1 迭代立项", "产品定义/复用与影响分析/计划"), ("P2 设计", "模块升级设计/投模/验证计划"), ("P3 EVT", "样机与专项验证"), ("P4 DVT", "模具件/可靠性/认证"), ("P5 PVT", "小批试产/发布评审"), ("P6 MP", "量产与关闭")]),
    (r"认证路线|认证前置|认证路径", [("整机安规", ""), ("EMC", ""), ("电芯/电池包认证", "IEC 62133/UL 等"), ("运输 UN38.3", ""), ("化学环保 RoHS/REACH", ""), ("标签/铭牌/说明书", "")]),
    (r"模具开发", [("模流分析", ""), ("模具设计评审", ""), ("开模加工", ""), ("T0 试模", ""), ("T1 试模", ""), ("T2/修模", ""), ("纹面与终验收", "")]),
    (r"产能爬坡", [("首批量产", ""), ("25% 产能", ""), ("50% 产能", ""), ("100% 产能", "")]),
    (r"切换|库存", [("旧版停投点确认", ""), ("在制品处理", ""), ("成品库存消耗策略", ""), ("物料切换", ""), ("文件/系统切换", ""), ("渠道与售后通知", "")]),
    (r"项目计划|甘特|WBS|时程", [("P1 立项/规划", ""), ("P2 设计", ""), ("P3 验证", ""), ("P4 试产", ""), ("P5 量产", "")]),
]


def plan_rows_for(name):
    for pat, rows in PLAN_ROW_SETS:
        if re.search(pat, name):
            return rows
    return []


# 认证归档行预置（F10）
CERT_ROW_SETS = [
    (r"UN38\.3", ["UN38.3 运输测试报告"]),
    (r"MSDS", ["MSDS 化学品安全说明书"]),
    (r"电芯|电池", ["电芯安全认证（IEC 62133/UL1642 等）", "电池包安全认证", "UN38.3 运输测试", "MSDS"]),
    (r"ECN|发布", ["ECN 变更通知", "受控文件发布记录"]),
    (r".", ["整机安规认证", "EMC 认证", "电池/电芯认证", "运输 UN38.3", "RoHS/REACH", "标签/铭牌核准"]),
]


def cert_rows_for(name):
    for pat, rows in CERT_ROW_SETS:
        if re.search(pat, name):
            return rows
    return []


# 复用确认边界项预置（F11）
BOUNDARY_SETS = [
    (r"电芯厂|资质", ["供应商与产线", "化学体系与型号", "批次一致性与来料记录", "变更通知机制", "审核报告有效期与范围"]),
    (r"保护电路", ["最大负载电流", "充电策略", "过充/过放/过流/过温/短路参数", "温升路径与散热", "固定方式与受压工况"]),
    (r"认证补测", ["目标市场", "标准版本", "产品差异点", "标签与说明书", "运输边界"]),
    (r"二供|定点", ["主供/二供型号一致性", "规格书与认证覆盖", "替代验证范围", "切换条件"]),
]
DEFAULT_BOUNDARIES = ["最大放电电流/负载", "充电策略", "温升路径/散热结构", "固定方式/结构工况", "目标市场/运输边界", "其他"]


def boundaries_for(name):
    for pat, rows in BOUNDARY_SETS:
        if re.search(pat, name):
            return rows
    return DEFAULT_BOUNDARIES


# ── 通用小构件 ──────────────────────────────────────────────────────────

def header_block(ws, ctx, ncols, note=None):
    """精简表头：标题 + 族标签 + （可选）一句话用法 + 项目信息行。取代大封面。"""
    last = get_column_letter(ncols)
    _merge(ws, f"A1:{last}1")
    t = ws.cell(row=1, column=1, value=ctx["title"])
    t.font = F_TITLE
    ws.row_dimensions[1].height = 30
    _merge(ws, f"A2:{last}2")
    ws.cell(row=2, column=1, value=f"{ctx['fid']} {FAMILIES[ctx['fid']]} · 适用：{ctx['applies']}").font = F_TAG
    row = 3
    if note:
        _merge(ws, f"A{row}:{last}{row}")
        n = ws.cell(row=row, column=1, value=note)
        n.font = F_HINT
        n.alignment = WRAP
        row += 1
    labels = ["项目编号", "项目名称", "负责人", "日期"]
    per = max(1, ncols // 4)
    col = 1
    for i, lab in enumerate(labels):
        end = ncols if i == 3 else min(col + per - 1, ncols)
        _cell(ws, f"{get_column_letter(col)}{row}", lab, font=F_SEC, fill=FILL_SEC)
        if end > col:
            _merge(ws, f"{get_column_letter(col + 1)}{row}:{get_column_letter(end)}{row}")
        for c in range(col, end + 1):
            ws.cell(row=row, column=c).border = BORDER
        col = end + 1
        if col > ncols:
            break
    ws.row_dimensions[row].height = 22
    return row + 2


def example_row(ws, row, values, height=20):
    """灰色斜体示例行：员工照着填，填的时候直接覆盖。合并单元格非锚点只画边框。"""
    ws.row_dimensions[row].height = height
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i)
        if v is not None and not isinstance(c, MergedCell):
            c.value = v
            c.font = F_EX
        c.border = BORDER
        c.alignment = WRAP


# ── 各族构建器（v2） ────────────────────────────────────────────────────

def build_f01(wb, ctx):
    ws = wb.active
    ws.title = "评审记录"
    ncols = 6
    _ws_defaults(ws, [14, 18, 14, 18, 12, 14])
    _merge(ws, "A1:F1")
    t = ws.cell(row=1, column=1, value=ctx["title"])
    t.font = F_TITLE
    ws.row_dimensions[1].height = 30
    _merge(ws, "A2:F2")
    ws.cell(row=2, column=1, value=f"{ctx['fid']} {FAMILIES[ctx['fid']]} · 适用：{ctx['applies']} · 可直接打印线下使用").font = F_TAG

    row = 4
    pairs = [("项目编号", "项目名称"), ("阶段 / Gate", "评审轮次"), ("评审日期", "主持人 / 记录人")]
    for a, b in pairs:
        _cell(ws, f"A{row}", a, font=F_SEC, fill=FILL_SEC)
        _merge(ws, f"B{row}:C{row}")
        for i in range(2, 4):
            ws.cell(row=row, column=i).border = BORDER
        _cell(ws, f"D{row}", b, font=F_SEC, fill=FILL_SEC)
        _merge(ws, f"E{row}:F{row}")
        for i in range(5, 7):
            ws.cell(row=row, column=i).border = BORDER
        ws.row_dimensions[row].height = 22
        row += 1
    _cell(ws, f"A{row}", "参会人员", font=F_SEC, fill=FILL_SEC)
    _merge(ws, f"B{row}:F{row}")
    for i in range(2, 7):
        ws.cell(row=row, column=i).border = BORDER
    ws.row_dimensions[row].height = 26
    row += 2

    row = _section(ws, row, ncols, "§1 入口检查", "逐条确认并写依据；不满足先补齐再评审。")
    _table_header(ws, row, ["✓", "检查项", "", "", "依据 / 说明", ""])
    ws.merge_cells(f"B{row}:D{row}")
    ws.merge_cells(f"E{row}:F{row}")
    row += 1
    for i in range(4):
        ws.merge_cells(f"B{row}:D{row}")
        ws.merge_cells(f"E{row}:F{row}")
        _blank_rows(ws, row, 1, ncols, height=22)
        row += 1
    row += 1

    row = _section(ws, row, ncols, "§2 评审材料与结论要点", "结论要点写「数字 + 决定」；判定为“有条件”的行，条件落到 §3。")
    _table_header(ws, row, ["序号", "材料 / 议题", "", "结论要点", "", "判定"])
    ws.merge_cells(f"B{row}:C{row}")
    ws.merge_cells(f"D{row}:E{row}")
    row += 1
    dv_start = row
    for i in range(6):
        ws.cell(row=row, column=1, value=i + 1).font = F_BODY
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=1).alignment = CENTER
        ws.merge_cells(f"B{row}:C{row}")
        ws.merge_cells(f"D{row}:E{row}")
        _blank_rows(ws, row, 1, ncols, height=24)
        row += 1
    _dv(ws, ["通过", "有条件", "不通过"], f"F{dv_start}:F{row - 1}")
    row += 1

    row = _section(ws, row, ncols, "§3 评审结论（三选一）", "有条件通过：条件+责任人+期限缺一不可；不通过：停留本阶段整改后重审，不回退。")
    _cell(ws, f"A{row}", "结论", font=F_SEC, fill=FILL_SEC)
    _merge(ws, f"B{row}:F{row}")
    for i in range(2, 7):
        ws.cell(row=row, column=i).border = BORDER
    _dv(ws, ["通过 approved", "有条件通过 conditional", "不通过 rejected"], f"B{row}")
    ws.row_dimensions[row].height = 22
    row += 1
    _table_header(ws, row, ["#", "条件项", "", "责任人", "期限", "关闭"])
    ws.merge_cells(f"B{row}:C{row}")
    row += 1
    ws.merge_cells(f"B{row}:C{row}")
    example_row(ws, row, [1, "示例：模具报价三家比价后回签（覆盖填写）", "", "李工", "07-22", ""], height=22)
    row += 1
    for i in range(2):
        ws.merge_cells(f"B{row}:C{row}")
        _blank_rows(ws, row, 1, ncols, height=22)
        ws.cell(row=row, column=1, value=i + 2).alignment = CENTER
        row += 1
    row += 1

    row = _signoff(ws, row, ncols, ("PM", "QA", "PE/MFG", "管理层/Owner"))
    _print_setup(ws, row, ncols)


def _test_sheet(wb, sheet_name, item_label):
    s = wb.create_sheet(sheet_name[:31])
    _ws_defaults(s, [14, 22, 22, 22, 14, 14])
    _merge(s, "A1:F1")
    s.cell(row=1, column=1, value=f"测试项：{item_label}").font = F_SEC
    r = 2
    for label, hint in [("判定标准", "量化标准，如：电芯表面 ≤ 70℃"), ("设备与环境", ""), ("样品编号", "")]:
        _cell(s, f"A{r}", label, font=F_SEC, fill=FILL_SEC)
        _merge(s, f"B{r}:F{r}")
        c = s.cell(row=r, column=2, value=hint)
        c.font = F_HINT
        for i in range(2, 7):
            s.cell(row=r, column=i).border = BORDER
        s.row_dimensions[r].height = 22
        r += 1
    r += 1
    _table_header(s, r, ["样品 SN", "条件/工况", "测量值", "标准", "判定", "备注"])
    r += 1
    dv0 = r
    example_row(s, r, ["#01", "常温满电", "68℃", "≤70℃", "Pass", "示例行，覆盖填写"])
    r += 1
    _blank_rows(s, r, 7, 6)
    r += 7
    _dv(s, ["Pass", "Fail", "NA"], f"E{dv0}:E{r - 1}")
    r += 1
    _cell(s, f"A{r}", "结论", font=F_SEC, fill=FILL_SEC)
    _merge(s, f"B{r}:F{r}")
    for i in range(2, 7):
        s.cell(row=r, column=i).border = BORDER
    s.row_dimensions[r].height = 26
    r += 1
    _cell(s, f"A{r}", "问题编号", font=F_SEC, fill=FILL_SEC)
    _merge(s, f"B{r}:F{r}")
    c = s.cell(row=r, column=2, value="Fail 必须开问题单并登记编号（联动「问题清单」sheet）")
    c.font = F_HINT
    for i in range(2, 7):
        s.cell(row=r, column=i).border = BORDER


def build_f02(wb, ctx):
    items = test_items_for(ctx["title"])
    ws = wb.active
    ws.title = "主页"
    _ws_defaults(ws, [16, 24, 16, 24])
    row = header_block(ws, ctx, 4, note="填法：每个测试项一个 sheet（已按常见项目建好，不需要的删掉、缺的复制「测试项-空白」）；填完回「目录」标判定。")
    for a, b in [("样品版本 / 数量", "软硬件配置"), ("测试周期", "测试负责人")]:
        _cell(ws, f"A{row}", a, font=F_SEC, fill=FILL_SEC)
        _cell(ws, f"B{row}", "")
        _cell(ws, f"C{row}", b, font=F_SEC, fill=FILL_SEC)
        _cell(ws, f"D{row}", "")
        ws.row_dimensions[row].height = 22
        row += 1
    _cell(ws, f"A{row}", "整体判定", font=F_SEC, fill=FILL_SEC)
    _merge(ws, f"B{row}:D{row}")
    for i in range(2, 5):
        ws.cell(row=row, column=i).border = BORDER
    _dv(ws, ["通过", "不通过（列未关闭问题编号）", "有条件"], f"B{row}")
    ws.row_dimensions[row].height = 24
    row += 2
    row = _signoff(ws, row, 4, ("测试", "QA", "批准"))

    toc = wb.create_sheet("目录")
    _ws_defaults(toc, [6, 26, 30, 18, 10])
    _table_header(toc, 1, ["序号", "测试项（对应同名 sheet）", "判定标准", "结果摘要", "判定"])
    r = 2
    for i, item in enumerate(items):
        toc.cell(row=r, column=1, value=i + 1).alignment = CENTER
        _blank_rows(toc, r, 1, 5, height=20)
        toc.cell(row=r, column=1, value=i + 1).border = BORDER
        c = toc.cell(row=r, column=2, value=item)
        c.font = F_BODY
        r += 1
    for _ in range(3):
        _blank_rows(toc, r, 1, 5, height=20)
        r += 1
    _dv(toc, ["Pass", "Fail", "NA"], f"E2:E{r - 1}")

    for item in items:
        _test_sheet(wb, item, item)
    _test_sheet(wb, "测试项-空白", "（复制本 sheet 后改名）")

    q = wb.create_sheet("问题清单")
    _ws_defaults(q, [10, 26, 26, 8, 14, 12, 10])
    _table_header(q, 1, ["编号", "现象 / 根因", "改善方案", "等级", "责任人", "期限", "状态"])
    example_row(q, 2, ["ISS-01", "示例：温升超标 4℃；进风口偏小", "风道口扩大，改模后复测", "P1", "周工", "09-30", "整改中"])
    _blank_rows(q, 3, 9, 7)
    _dv(q, ["P0", "P1", "P2", "P3"], "D2:D11")
    _dv(q, ["打开", "整改中", "已关闭", "已豁免"], "G2:G11")


def build_f03(wb, ctx):
    ws = wb.active
    ws.title = "清单"
    _ws_defaults(ws, [10, 14, 26, 24, 24, 8, 12, 12, 10])
    row = header_block(ws, ctx, 9, note="一行一问题：现象+根因+对策+责任人期限缺一不可；P0/P1 未关闭不得过 Gate。")
    _table_header(ws, row, ["编号", "来源/阶段", "现象描述", "根因分析", "对策/改善方案", "等级", "责任人", "期限", "状态"])
    row += 1
    dv0 = row
    example_row(ws, row, ["ISS-01", "EVT", "示例：气压精度超差 ±0.3bar", "传感器批次漂移", "来料加抽检，EOL 校准点+1", "P1", "陈工", "09-30", "整改中"])
    row += 1
    _blank_rows(ws, row, 14, 9, height=22)
    row += 14
    _dv(ws, ["P0", "P1", "P2", "P3"], f"F{dv0}:F{row - 1}")
    _dv(ws, ["打开", "整改中", "已关闭", "已豁免"], f"I{dv0}:I{row - 1}")


def build_f04(wb, ctx):
    ws = wb.active
    ws.title = "影响矩阵"
    _ws_defaults(ws, [16, 10, 30, 10, 26, 12])
    row = header_block(ws, ctx, 6, note="逐域判断是否受影响；受影响的必须有应对措施和责任人。")
    _table_header(ws, row, ["影响域", "是否受影响", "影响说明", "风险等级", "应对措施 / 验证项", "责任人"])
    row += 1
    domains = ["产品定义/规格", "结构 / 模具", "PCBA / 电源", "软件 / 固件", "电池 / 电芯 / 安全",
               "BOM / 成本", "供应商 / 交期", "认证 / 合规", "包装 / 标签 / 运输", "产线 / 治具 / EOL",
               "库存 / 在制品", "售后 / 市场"]
    dv0 = row
    for d in domains:
        _blank_rows(ws, row, 1, 6, height=22)
        _cell(ws, f"A{row}", d, font=F_SEC, fill=FILL_SEC)
        row += 1
    _dv(ws, ["是", "否"], f"B{dv0}:B{row - 1}")
    _dv(ws, ["高", "中", "低"], f"D{dv0}:D{row - 1}")
    row += 1
    _cell(ws, f"A{row}", "总体结论", font=F_SEC, fill=FILL_SEC)
    _merge(ws, f"B{row}:F{row}")
    for i in range(2, 7):
        ws.cell(row=row, column=i).border = BORDER
    _dv(ws, ["影响可控", "需专项验证", "建议转轨（ECO/DRV/NPD）"], f"B{row}")
    ws.row_dimensions[row].height = 24

    if re.search(r"复用|FMEA|CTQ", ctx["title"]):
        s = wb.create_sheet("模块复用策略" if "复用" in ctx["title"] else "失效模式")
        if "复用" in ctx["title"]:
            _ws_defaults(s, [20, 16, 30, 30])
            _table_header(s, 1, ["模块", "复用等级", "依据（版本/认证/测试记录）", "边界与风险"])
            mods = ["电池/电芯/电池包", "机芯/马达/泵体", "PCBA/电源/主控", "软件/固件/APP接口", "结构/外壳/模具", "包装/标签/认证边界"]
            for i, m in enumerate(mods):
                r = 2 + i
                _blank_rows(s, r, 1, 4, height=24)
                _cell(s, f"A{r}", m, font=F_SEC, fill=FILL_SEC)
            _dv(s, ["直接复用", "复用+适配验证", "轻量修改", "重新开发"], "B2:B7")
        else:
            _ws_defaults(s, [18, 20, 20, 8, 8, 8, 10, 22, 12])
            _table_header(s, 1, ["部件/工序", "失效模式", "失效影响", "S", "O", "D", "RPN", "控制/改进措施", "责任人"])
            example_row(s, 2, ["电池仓", "跌落后固定失效", "电芯位移/短路", 8, 3, 4, 96, "加限位筋+跌落验证", "结构"])
            _blank_rows(s, 3, 12, 9)


def build_f05(wb, ctx):
    ws = wb.active
    ws.title = "BOM表"
    _ws_defaults(ws, [8, 16, 20, 24, 8, 8, 16, 10, 12, 10, 16])
    row = header_block(ws, ctx, 11, note="状态列用下拉（新增/替换/停用/复用）；金额自动计算，「差异汇总」自动计数。")
    _table_header(ws, row, ["阶层", "料号", "品名", "规格", "用量", "单位", "供应商", "单价", "金额", "状态", "备注"])
    row += 1
    example_row(ws, row, [1, "BP-CELL-21700", "电芯", "21700 4000mAh", 2, "PCS", "供应商A", 12.5, "", "复用", "示例行，覆盖填写"], height=20)
    ws.cell(row=row, column=9, value=f'=IF(OR(E{row}="",H{row}=""),"",E{row}*H{row})').font = F_EX
    row += 1
    n = 34
    dv0 = row - 1
    for i in range(n):
        _blank_rows(ws, row, 1, 11, height=18)
        ws.cell(row=row, column=9, value=f'=IF(OR(E{row}="",H{row}=""),"",E{row}*H{row})').font = F_BODY
        row += 1
    _dv(ws, ["新增", "替换", "停用", "复用"], f"J{dv0}:J{row - 1}")
    _cell(ws, f"H{row}", "合计", font=F_SEC, fill=FILL_SEC, align=CENTER)
    ws.cell(row=row, column=9, value=f"=SUM(I{dv0}:I{row - 1})").font = F_SEC
    ws.cell(row=row, column=9).border = BORDER

    s = wb.create_sheet("差异汇总")
    _ws_defaults(s, [16, 12, 40])
    _table_header(s, 1, ["状态", "行数", "说明"])
    for i, st in enumerate(["新增", "替换", "停用", "复用"]):
        r = 2 + i
        _cell(s, f"A{r}", st, font=F_SEC, fill=FILL_SEC)
        s.cell(row=r, column=2, value=f'=COUNTIF(BOM表!J:J,"{st}")').font = F_BODY
        s.cell(row=r, column=2).border = BORDER
        _cell(s, f"C{r}", "")


def build_f06(wb, ctx):
    ws = wb.active
    ws.title = "计划"
    _ws_defaults(ws, [16, 26, 12, 12, 12, 12, 10])
    row = header_block(ws, ctx, 7, note="按预置行填负责人和日期即可；行不够往下加。系统内有自动排期，本表用于冻结基线与线下沟通。")
    _table_header(ws, row, ["阶段/类别", "任务 / 里程碑", "负责人", "开始", "结束", "交付物", "状态"])
    row += 1
    prefill = plan_rows_for(ctx["title"])
    dv0 = row
    for a, b in prefill:
        _blank_rows(ws, row, 1, 7, height=20)
        _cell(ws, f"A{row}", a, font=F_SEC, fill=FILL_SEC)
        c = ws.cell(row=row, column=2, value=b)
        c.font = F_HINT if b else F_BODY
        c.border = BORDER
        row += 1
    extra = 6 if prefill else 14
    _blank_rows(ws, row, extra, 7, height=20)
    row += extra
    _dv(ws, ["待开始", "进行中", "已完成", "已跳过"], f"G{dv0}:G{row - 1}")
    row += 1
    _cell(ws, f"A{row}", "关键风险", font=F_SEC, fill=FILL_SEC)
    _merge(ws, f"B{row}:G{row}")
    for i in range(2, 8):
        ws.cell(row=row, column=i).border = BORDER
    ws.row_dimensions[row].height = 30


def build_f07(wb, ctx):
    ws = wb.active
    ws.title = "清单"
    _ws_defaults(ws, [8, 30, 12, 12, 14, 12, 20])
    row = header_block(ws, ctx, 7, note="逐行登记条目与状态；作为 Gate 证据时以「已确认」为准。")
    _table_header(ws, row, ["序号", "条目 / 文件名", "版本", "日期", "负责人", "状态", "存放位置 / 链接"])
    row += 1
    prefill = list_items_for(ctx["title"])
    dv0 = row
    for i, item in enumerate(prefill):
        _blank_rows(ws, row, 1, 7, height=20)
        ws.cell(row=row, column=1, value=i + 1).alignment = CENTER
        ws.cell(row=row, column=1).border = BORDER
        c = ws.cell(row=row, column=2, value=item)
        c.font = F_BODY
        row += 1
    extra = 5 if prefill else 12
    for i in range(extra):
        _blank_rows(ws, row, 1, 7, height=20)
        ws.cell(row=row, column=1, value=len(prefill) + i + 1).alignment = CENTER
        ws.cell(row=row, column=1).border = BORDER
        row += 1
    _dv(ws, ["待提交", "已提交", "已确认", "不适用"], f"F{dv0}:F{row - 1}")


def build_f08(wb, ctx):
    ws = wb.active
    ws.title = "试产与良率"
    _ws_defaults(ws, [20, 10, 10, 10, 34])
    row = header_block(ws, ctx, 5, note="FPY/RTY 自动计算；关键异常填「问题关闭」sheet，发布看「就绪确认」。")
    for a, b in [("试产批量", "产线 / 班组"), ("试产日期", "目标良率")]:
        _cell(ws, f"A{row}", a, font=F_SEC, fill=FILL_SEC)
        _cell(ws, f"B{row}", "")
        _cell(ws, f"C{row}", b, font=F_SEC, fill=FILL_SEC)
        _merge(ws, f"D{row}:E{row}")
        for i in range(4, 6):
            ws.cell(row=row, column=i).border = BORDER
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1
    _table_header(ws, row, ["工位", "投入", "不良", "FPY", "TOP 不良现象"])
    row += 1
    stations = ["SMT / PCBA", "半成品组装", "整机组装", "整机 EOL 100% 测试", "包装"]
    first = row
    for st in stations:
        _blank_rows(ws, row, 1, 5, height=20)
        _cell(ws, f"A{row}", st, font=F_SEC, fill=FILL_SEC)
        ws.cell(row=row, column=4, value=f'=IF(N(B{row})=0,"",1-C{row}/B{row})').number_format = "0.0%"
        ws.cell(row=row, column=4).font = F_BODY
        row += 1
    _cell(ws, f"A{row}", "直通率 RTY", font=F_SEC, fill=FILL_SEC)
    ws.cell(row=row, column=4, value=f'=IF(COUNT(D{first}:D{row - 1})=0,"",PRODUCT(D{first}:D{row - 1}))').number_format = "0.0%"
    ws.cell(row=row, column=4).font = F_SEC
    for c in (2, 3, 4, 5):
        ws.cell(row=row, column=c).border = BORDER

    q = wb.create_sheet("问题关闭")
    _ws_defaults(q, [10, 14, 26, 10, 24, 12, 12])
    _table_header(q, 1, ["编号", "工位", "不良现象 / 根因", "数量", "对策", "验证结果", "状态"])
    example_row(q, 2, ["ISS-01", "组装", "示例：风道卡扣断裂；模内应力集中", 4, "修模加 R 角，装配治具限位", "复测通过", "已关闭"])
    _blank_rows(q, 3, 9, 7)
    _dv(q, ["已关闭", "整改中", "打开"], "G2:G11")

    rd = wb.create_sheet("发布就绪确认")
    _ws_defaults(rd, [6, 44, 20, 14])
    _table_header(rd, 1, ["✓", "确认项", "证据编号", "确认人"])
    items = ["SOP/WI 与检验标准已按试产问题更新", "治具 / EOL 测试程序验收通过，EOL 覆盖 100% 出货单元",
             "认证证据归档（UN38.3 / MSDS / 电芯认证）编号已登记", "版本切换与库存处理方案已签核",
             "P0/P1 问题全部关闭"]
    for i, it in enumerate(items):
        r = 2 + i
        _blank_rows(rd, r, 1, 4, height=22)
        rd.cell(row=r, column=2, value=it).font = F_BODY


def build_f09(wb, ctx):
    ws = wb.active
    ws.title = "工序步骤"
    _ws_defaults(ws, [10, 8, 30, 26, 16, 14])
    row = header_block(ws, ctx, 6, note="一行一步骤；关键要点写防错方法；图示可直接插入单元格区域。")
    _table_header(ws, row, ["工位", "步骤", "作业内容", "关键要点 / 防错", "检验项", "工具 / 治具"])
    row += 1
    example_row(ws, row, ["组装-1", 1, "示例：装机芯入仓，卡扣到位", "听到咔哒声；不可斜插", "卡扣目检", "定位治具 J-01"], height=22)
    row += 1
    _blank_rows(ws, row, 14, 6, height=22)

    s = wb.create_sheet("检验标准")
    _ws_defaults(s, [20, 30, 16, 14, 12])
    _table_header(s, 1, ["检验项目", "标准（量化）", "方法 / 设备", "抽样方案", "判定"])
    example_row(s, 2, ["气压精度", "示例：±0.1 bar", "标准压力计比对", "每批 5 台", "全部合格"])
    _blank_rows(s, 3, 9, 5)

    v = wb.create_sheet("变更记录")
    _ws_defaults(v, [10, 12, 40, 14])
    _table_header(v, 1, ["版本", "日期", "变更内容（对应问题编号）", "批准人"])
    _blank_rows(v, 2, 5, 4)


def build_f10(wb, ctx):
    ws = wb.active
    ws.title = "证据索引"
    _ws_defaults(ws, [24, 16, 14, 12, 12, 20, 16, 12])
    row = header_block(ws, ctx, 8, note="一行一份证书/报告；复用旧证据的行状态选「复用确认」并填「复用确认」sheet。缺证据不得发布 MP。")
    _table_header(ws, row, ["证书 / 报告名称", "编号", "机构", "签发日", "有效期", "适用范围（型号/电芯）", "文件位置", "状态"])
    row += 1
    dv0 = row
    for name in cert_rows_for(ctx["title"]):
        _blank_rows(ws, row, 1, 8, height=20)
        c = ws.cell(row=row, column=1, value=name)
        c.font = F_BODY
        row += 1
    _blank_rows(ws, row, 4, 8, height=20)
    row += 4
    _dv(ws, ["有效", "待补测", "已过期", "复用确认"], f"H{dv0}:H{row - 1}")

    s = wb.create_sheet("复用确认")
    _ws_defaults(s, [22, 26, 26, 12])
    _table_header(s, 1, ["边界项", "原批准范围", "本项目用法", "是否超界"])
    for i, b in enumerate(["型号 / 电芯", "目标市场 / 运输边界", "标签 / 说明书", "充电策略 / 温升路径"]):
        r = 2 + i
        _blank_rows(s, r, 1, 4, height=22)
        _cell(s, f"A{r}", b, font=F_SEC, fill=FILL_SEC)
    _dv(s, ["未超界", "超界"], "D2:D5")


def build_f11(wb, ctx):
    ws = wb.active
    ws.title = "复用确认"
    ncols = 5
    _ws_defaults(ws, [18, 22, 22, 12, 16])
    _merge(ws, "A1:E1")
    ws.cell(row=1, column=1, value=ctx["title"]).font = F_TITLE
    ws.row_dimensions[1].height = 30
    _merge(ws, "A2:E2")
    ws.cell(row=2, column=1, value=f"{ctx['fid']} {FAMILIES[ctx['fid']]} · 适用：{ctx['applies']} · 全部「未超界」时本单可替代全量审核/重新认证").font = F_TAG
    row = 4
    for a, b in [("项目编号", "复用对象"), ("原批准载体", "复用等级")]:
        _cell(ws, f"A{row}", a, font=F_SEC, fill=FILL_SEC)
        _cell(ws, f"B{row}", "")
        _cell(ws, f"C{row}", b, font=F_SEC, fill=FILL_SEC)
        _merge(ws, f"D{row}:E{row}")
        for i in range(4, 6):
            ws.cell(row=row, column=i).border = BORDER
        ws.row_dimensions[row].height = 22
        row += 1
    _dv(ws, ["直接复用", "复用+适配验证", "轻量修改", "重新开发"], f"D{row - 1}")
    row += 1
    row = _section(ws, row, ncols, "§1 边界对比（每行回答：这次用法有没有超出原批准范围）")
    _table_header(ws, row, ["边界项", "原批准范围", "本项目用法", "是否超界", "备注"])
    row += 1
    dv0 = row
    for b in boundaries_for(ctx["title"]):
        _blank_rows(ws, row, 1, ncols, height=22)
        _cell(ws, f"A{row}", b, font=F_SEC, fill=FILL_SEC)
        row += 1
    _dv(ws, ["未超界", "超界"], f"D{dv0}:D{row - 1}")
    row += 1
    row = _section(ws, row, ncols, "§2 超界项处理（任何一行超界都要有补做动作）")
    _table_header(ws, row, ["超界项", "补做验证 / 认证", "", "责任人", "期限"])
    ws.merge_cells(f"B{row}:C{row}")
    row += 1
    for _ in range(3):
        ws.merge_cells(f"B{row}:C{row}")
        _blank_rows(ws, row, 1, ncols, height=22)
        row += 1
    row += 1
    row = _section(ws, row, ncols, "§3 结论与证据编号")
    _merge(ws, f"A{row}:E{row}")
    for i in range(1, 6):
        ws.cell(row=row, column=i).border = BORDER
    ws.row_dimensions[row].height = 30
    row += 2
    row = _signoff(ws, row, ncols, ("电池安全/技术", "QA", "SCM"))
    _print_setup(ws, row, ncols)


def build_f12(wb, ctx):
    ws = wb.active
    ws.title = "客户签核"
    ncols = 4
    _ws_defaults(ws, [18, 26, 18, 26])
    _merge(ws, "A1:D1")
    ws.cell(row=1, column=1, value=ctx["title"]).font = F_TITLE
    ws.row_dimensions[1].height = 30
    _merge(ws, "A2:D2")
    ws.cell(row=2, column=1, value=f"{ctx['fid']} Customer Sign-off · 适用：{ctx['applies']} · 可打印线下签署").font = F_TAG
    row = 4
    for a, b in [("项目编号 Project", "客户 Customer"), ("产品型号 Model", "版本 Version"), ("签核对象 Subject", "日期 Date")]:
        _cell(ws, f"A{row}", a, font=F_SEC, fill=FILL_SEC)
        _cell(ws, f"B{row}", "")
        _cell(ws, f"C{row}", b, font=F_SEC, fill=FILL_SEC)
        _cell(ws, f"D{row}", "")
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1
    row = _section(ws, row, ncols, "签核内容 Description", "写清签核的样品/文件/规格版本与判定口径。")
    _merge(ws, f"A{row}:D{row + 2}")
    for r in range(row, row + 3):
        for i in range(1, 5):
            ws.cell(row=r, column=i).border = BORDER
    ws.row_dimensions[row].height = 24
    row += 4
    row = _section(ws, row, ncols, "样品 / 文件清单 Items")
    _table_header(ws, row, ["#", "名称 Item", "版本/SN", "备注 Note"])
    row += 1
    for i in range(4):
        _blank_rows(ws, row, 1, 4, height=20)
        ws.cell(row=row, column=1, value=i + 1).alignment = CENTER
        ws.cell(row=row, column=1).border = BORDER
        row += 1
    row += 1
    row = _section(ws, row, ncols, "客户判定 Decision")
    _cell(ws, f"A{row}", "判定 Decision", font=F_SEC, fill=FILL_SEC)
    _merge(ws, f"B{row}:D{row}")
    for i in range(2, 5):
        ws.cell(row=row, column=i).border = BORDER
    _dv(ws, ["接受 Accepted", "有条件接受 Conditional", "拒绝 Rejected"], f"B{row}")
    ws.row_dimensions[row].height = 22
    row += 1
    _cell(ws, f"A{row}", "条件 Conditions", font=F_SEC, fill=FILL_SEC)
    _merge(ws, f"B{row}:D{row}")
    for i in range(2, 5):
        ws.cell(row=row, column=i).border = BORDER
    ws.row_dimensions[row].height = 30
    row += 2
    row = _section(ws, row, ncols, "双方签署 Signatures")
    _cell(ws, f"A{row}", "客户代表\nCustomer", font=F_SEC, fill=FILL_SEC)
    _cell(ws, f"B{row}", "")
    _cell(ws, f"C{row}", "制造方代表\nManufacturer", font=F_SEC, fill=FILL_SEC)
    _cell(ws, f"D{row}", "")
    ws.row_dimensions[row].height = 34
    row += 1
    _cell(ws, f"A{row}", "日期 Date", font=F_SEC, fill=FILL_SEC)
    _cell(ws, f"B{row}", "")
    _cell(ws, f"C{row}", "日期 Date", font=F_SEC, fill=FILL_SEC)
    _cell(ws, f"D{row}", "")
    row += 1
    _print_setup(ws, row, ncols)


def build_f13(wb, ctx):
    ws = wb.active
    ws.title = "文档"
    ncols = 4
    _ws_defaults(ws, [22, 30, 30, 30])
    row = header_block(ws, ctx, ncols, note="按预置章节直接往格子里写；表格数据可直接粘贴到对应区域。")
    for title, hint in doc_sections_for(ctx["title"]):
        _merge(ws, f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.font = F_SEC
        c.fill = FILL_SEC
        for i in range(1, ncols + 1):
            ws.cell(row=row, column=i).border = BORDER
            ws.cell(row=row, column=i).fill = FILL_SEC
        ws.row_dimensions[row].height = 20
        row += 1
        _merge(ws, f"A{row}:{get_column_letter(ncols)}{row + 1}")
        h = ws.cell(row=row, column=1, value=hint)
        h.font = F_HINT
        h.alignment = WRAP
        for r in range(row, row + 2):
            for i in range(1, ncols + 1):
                ws.cell(row=r, column=i).border = BORDER
            ws.row_dimensions[r].height = 34
        row += 2
    row += 1
    _signoff(ws, row, ncols)


def build_f14(wb, ctx):
    ws = wb.active
    ws.title = "样机记录"
    _ws_defaults(ws, [8, 16, 22, 20, 12, 16])
    row = header_block(ws, ctx, 6, note="上半部分记一次 Build 的配置基线，下表逐台登记 SN 与去向。")
    for a, b in [("Build 版本 / 数量", "BOM 版本"), ("软件/固件版本", "生产日期 / 制作人"), ("关键配置差异", "用途（EVT/DVT/PVT/首件）")]:
        _cell(ws, f"A{row}", a, font=F_SEC, fill=FILL_SEC)
        _merge(ws, f"B{row}:C{row}")
        for i in range(2, 4):
            ws.cell(row=row, column=i).border = BORDER
        _cell(ws, f"D{row}", b, font=F_SEC, fill=FILL_SEC)
        _merge(ws, f"E{row}:F{row}")
        for i in range(5, 7):
            ws.cell(row=row, column=i).border = BORDER
        ws.row_dimensions[row].height = 22
        row += 1
    row += 1
    _table_header(ws, row, ["#", "样机 SN", "配置差异", "去向（测试/客户/留样）", "状态", "异常与处置"])
    row += 1
    dv0 = row
    example_row(ws, row, [1, "EVT-B1-001", "示例：无差异", "可靠性测试", "在用", ""])
    row += 1
    for i in range(11):
        _blank_rows(ws, row, 1, 6, height=20)
        ws.cell(row=row, column=1, value=i + 2).alignment = CENTER
        ws.cell(row=row, column=1).border = BORDER
        row += 1
    _dv(ws, ["在用", "损坏", "已返还", "留样"], f"E{dv0}:E{row - 1}")


BUILDERS = {
    "F01": build_f01, "F02": build_f02, "F03": build_f03, "F04": build_f04,
    "F05": build_f05, "F06": build_f06, "F07": build_f07, "F08": build_f08,
    "F09": build_f09, "F10": build_f10, "F11": build_f11, "F12": build_f12,
    "F13": build_f13, "F14": build_f14,
}


def load_names(tsv_path):
    out = []
    with open(tsv_path, encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line or line.startswith("TOTAL"):
                continue
            parts = line.split("\t")
            name = parts[0].strip()
            occ = parts[1].split(",") if len(parts) > 1 else []
            out.append((name, occ))
    return out


def applies_text(occ):
    cats = []
    for tag in occ:
        cat = tag.split(":")[0].upper()
        if cat == "DERIVATIVE":
            cat = "DRV"
        if cat not in cats:
            cats.append(cat)
    tags = sorted({t.split(":")[0].upper().replace("DERIVATIVE", "DRV") + "·" + t.split(":")[1] for t in occ if ":" in t})
    return "、".join(tags[:8]) + ("…" if len(tags) > 8 else "")


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    tsv = args[0] if args else os.path.join(ROOT, "docs", "templates", "deliverables.tsv")
    items = load_names(tsv)
    dry = "--dry" in sys.argv

    mapping = []
    for name, occ in items:
        fid = classify(name)
        mapping.append((name, fid, occ))

    if dry:
        from collections import Counter
        cnt = Counter(f for _, f, _ in mapping)
        for fid in sorted(FAMILIES):
            print(f"\n== {fid} {FAMILIES[fid]} ({cnt.get(fid, 0)}) ==")
            for name, f, _ in mapping:
                if f == fid:
                    print("  " + name)
        return

    os.makedirs(OUT_ROOT, exist_ok=True)
    index_rows = []
    for name, fid, occ in mapping:
        folder = os.path.join(OUT_ROOT, f"{fid}-{FAMILIES[fid]}")
        os.makedirs(folder, exist_ok=True)
        fname = sanitize(name) + ".xlsx"
        ctx = {"title": name, "fid": fid, "applies": applies_text(occ)}
        template_path = os.path.join(folder, fname)
        wb = Workbook()
        BUILDERS[fid](wb, ctx)
        wb.save(template_path)
        index_rows.append((name, fid, FAMILIES[fid], f"{fid}-{FAMILIES[fid]}/{fname}", "、".join(occ)))

    # 索引工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "模板索引"
    _ws_defaults(ws, [34, 8, 18, 52, 40])
    _merge(ws, "A1:E1")
    ws.cell(row=1, column=1, value="交付物模板索引 · 共 %d 项" % len(index_rows)).font = F_TITLE
    ws.row_dimensions[1].height = 30
    _table_header(ws, 2, ["交付物名称", "族", "模板族", "文件路径", "出现位置（类型:阶段）"])
    for i, (name, fid, fam, path, occ) in enumerate(sorted(index_rows, key=lambda x: (x[1], x[0]))):
        r = 3 + i
        for col, val in enumerate([name, fid, fam, path, occ], start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = F_BODY
            c.border = BORDER
            c.alignment = WRAP
    ws.freeze_panes = "A3"
    wb.save(os.path.join(OUT_ROOT, "..", "模板索引.xlsx"))

    # shared 常量表（前后端共用：Gate 面板/任务交付物区「参照模板」入口 + 服务端下载端点白名单）
    ts_path = os.path.join(ROOT, "shared", "deliverable-templates.ts")
    lines = [
        "// GENERATED by scripts/generate-deliverable-templates.py — 不要手改，改生成器后重跑。",
        "// 交付物名称 → docs/templates/deliverables/ 下的模板文件相对路径。",
        "export const DELIVERABLE_TEMPLATE_FILES: Record<string, string> = {",
    ]
    for name, fid, fam, rel, _ in sorted(index_rows, key=lambda x: (x[1], x[0])):
        k = name.replace("\\", "\\\\").replace('"', '\\"')
        v = rel.replace("\\", "\\\\").replace('"', '\\"')
        lines.append(f'  "{k}": "{v}",')
    lines += [
        "};",
        "",
        "/** 交付物名称 → 模板相对路径；未知名称返回 null（服务端以此为白名单，杜绝任意路径）。 */",
        "export function getDeliverableTemplatePath(name: string): string | null {",
        "  return DELIVERABLE_TEMPLATE_FILES[name] ?? null;",
        "}",
        "",
    ]
    with open(ts_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"generated {len(index_rows)} templates + 模板索引.xlsx + shared/deliverable-templates.ts")


if __name__ == "__main__":
    main()
